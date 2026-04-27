from __future__ import annotations

import secrets
import string
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.core.security import hash_password
from backend.database.models import RevueDePresse
from backend.repositories.user_repo import UserRepository
from backend.schemas.user_schema import (
	UserActionResult,
	UserCreateRequest,
	UserImportResult,
	UserPasswordUpdateRequest,
	UserResponse,
	UserStatusUpdateRequest,
)
from backend.services.email_service import EmailService


class UserService:
	def __init__(self, user_repo: UserRepository | None = None) -> None:
		self.user_repo = user_repo or UserRepository()
		self.email_service = EmailService()

	def list_users(self, db: Session) -> list[UserResponse]:
		users = self.user_repo.list_users(db)
		return [self._to_user_response(user) for user in users]

	def create_user(self, db: Session, payload: UserCreateRequest) -> UserResponse:
		if self.user_repo.get_by_email(db, payload.email):
			raise HTTPException(
				status_code=status.HTTP_409_CONFLICT,
				detail="Email already registered",
			)

		username = payload.username or self._default_username(payload.email)
		if self.user_repo.get_by_username(db, username):
			raise HTTPException(
				status_code=status.HTTP_409_CONFLICT,
				detail="Username already taken",
			)

		plain_password = payload.password or self._generate_temporary_password()
		user = self.user_repo.create(
			db,
			email=payload.email,
			password_hash=hash_password(plain_password),
			username=username,
			role=self._normalize_role(payload.role),
		)
		user.is_active = payload.is_active
		db.commit()
		db.refresh(user)

		email_sent = self.email_service.send_account_created_email(
			to_email=user.email,
			username=user.username,
			role=user.role,
			plain_password=plain_password,
		)

		return self._to_user_response(
			user,
			email_sent=email_sent,
			generated_password=None if payload.password else plain_password,
		)

	def delete_user(self, db: Session, user_id: int) -> UserActionResult:
		user = self.user_repo.get_by_id(db, user_id)
		if user is None:
			raise HTTPException(
				status_code=status.HTTP_404_NOT_FOUND,
				detail="User not found",
			)

		revues_count = db.query(RevueDePresse).filter(RevueDePresse.user_id == user_id).count()
		if revues_count:
			raise HTTPException(
				status_code=status.HTTP_409_CONFLICT,
				detail="Cannot delete a user linked to published revues",
			)

		username = user.username
		self.user_repo.delete(db, user)
		db.commit()
		return UserActionResult(message=f"User '{username}' deleted successfully.")

	def update_user_password(self, db: Session, user_id: int, payload: UserPasswordUpdateRequest) -> UserActionResult:
		user = self.user_repo.get_by_id(db, user_id)
		if user is None:
			raise HTTPException(
				status_code=status.HTTP_404_NOT_FOUND,
				detail="User not found",
			)

		user.password_hash = hash_password(payload.password)
		self.user_repo.save(db, user)
		return UserActionResult(
			message="Password updated successfully.",
			user=self._to_user_response(user),
		)

	def update_user_status(self, db: Session, user_id: int, payload: UserStatusUpdateRequest) -> UserActionResult:
		user = self.user_repo.get_by_id(db, user_id)
		if user is None:
			raise HTTPException(
				status_code=status.HTTP_404_NOT_FOUND,
				detail="User not found",
			)

		user.is_active = payload.is_active
		self.user_repo.save(db, user)
		message = "User activated successfully." if payload.is_active else "User deactivated successfully."
		return UserActionResult(
			message=message,
			user=self._to_user_response(user),
		)

	def import_users_from_excel(self, db: Session, file_bytes: bytes, filename: str) -> UserImportResult:
		if not filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
			raise HTTPException(
				status_code=status.HTTP_400_BAD_REQUEST,
				detail="Invalid file type. Please upload an Excel .xlsx file.",
			)

		try:
			workbook = load_workbook(BytesIO(file_bytes), data_only=True)
		except Exception as exc:
			raise HTTPException(
				status_code=status.HTTP_400_BAD_REQUEST,
				detail=f"Unable to read Excel file: {exc}",
			)

		sheet = workbook.active
		rows = list(sheet.iter_rows(values_only=True))
		if len(rows) < 2:
			raise HTTPException(
				status_code=status.HTTP_400_BAD_REQUEST,
				detail="Excel file is empty or missing headers.",
			)

		header_row_index, headers = self._find_header_row(rows)
		required_headers = {"name", "email", "role"}
		if header_row_index is None or not required_headers.issubset(headers.keys()):
			raise HTTPException(
				status_code=status.HTTP_400_BAD_REQUEST,
				detail="Excel file must include columns: name/nom, email, role/rôle.",
			)

		parsed_users: list[dict[str, object]] = []
		errors: list[str] = []
		emails_seen: set[str] = set()
		usernames_seen: set[str] = set()

		for row_number, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
			if self._is_row_empty(row):
				continue

			row_data = self._extract_row_data(headers, row)
			row_errors = self._validate_row(row_data, row_number)
			if row_errors:
				errors.extend(row_errors)
				continue

			email = str(row_data["email"]).strip().lower()
			username = str(row_data["name"]).strip()
			role = self._normalize_role(str(row_data["role"]).strip())
			password = str(row_data.get("password") or "").strip() or None
			is_active = self._parse_bool(row_data.get("is_active"), default=True)

			if email in emails_seen:
				parsed_users.append(
					{
						"email": email,
						"username": username,
						"role": role,
						"password": password,
						"password_was_generated": password is None,
						"is_active": is_active,
						"skip_reason": "duplicate_email_in_file",
					},
				)
				continue
			if username in usernames_seen:
				parsed_users.append(
					{
						"email": email,
						"username": username,
						"role": role,
						"password": password,
						"password_was_generated": password is None,
						"is_active": is_active,
						"skip_reason": "duplicate_username_in_file",
					},
				)
				continue

			emails_seen.add(email)
			usernames_seen.add(username)
			parsed_users.append(
				{
					"email": email,
					"username": username,
					"role": role,
					"password": password,
					"password_was_generated": password is None,
					"is_active": is_active,
				},
			)

		if len(parsed_users) < 1:
			raise HTTPException(
				status_code=status.HTTP_400_BAD_REQUEST,
				detail="Excel file must contain at least 1 valid user.",
			)

		created_users: list[tuple[object, str]] = []
		skipped_count = 0
		try:
			for item in parsed_users:
				if item.get("skip_reason"):
					skipped_count += 1
					continue

				existing_email = self.user_repo.get_by_email(db, str(item["email"]))
				if existing_email is not None:
					skipped_count += 1
					continue

				existing_username = self.user_repo.get_by_username(db, str(item["username"]))
				if existing_username is not None:
					skipped_count += 1
					continue

				plain_password = str(item["password"]) if item["password"] else self._generate_temporary_password()
				user = self.user_repo.create(
					db,
					email=str(item["email"]),
					password_hash=hash_password(plain_password),
					username=str(item["username"]),
					role=str(item["role"]),
					commit=False,
				)
				user.is_active = bool(item["is_active"])
				created_users.append((user, plain_password))

			db.commit()
			for user, _ in created_users:
				db.refresh(user)
		except Exception:
			db.rollback()
			raise

		results: list[UserResponse] = []
		email_sent_count = 0
		for user, plain_password in created_users:
			email_sent = self.email_service.send_account_created_email(
				to_email=user.email,
				username=user.username,
				role=user.role,
				plain_password=plain_password,
			)
			if email_sent:
				email_sent_count += 1
			results.append(
				self._to_user_response(
					user,
					email_sent=email_sent,
					generated_password=plain_password if any(item for item in parsed_users if str(item["email"]) == user.email and item.get("password_was_generated")) else None,
				),
			)

		return UserImportResult(
			imported_count=len(results),
			skipped_count=skipped_count,
			email_sent_count=email_sent_count,
			users=results,
			message="Users imported successfully." if not errors else f"Users imported with warnings: {len(errors)} invalid row(s) ignored.",
		)

	@staticmethod
	def _to_user_response(
		user,
		*,
		email_sent: bool = False,
		generated_password: str | None = None,
	) -> UserResponse:
		return UserResponse(
			id=user.id,
			username=user.username,
			email=user.email,
			role=user.role,
			is_active=user.is_active,
			created_at=user.created_at,
			last_login=user.last_login,
			email_sent=email_sent,
			generated_password=generated_password,
		)

	@staticmethod
	def _generate_temporary_password(length: int = 10) -> str:
		alphabet = string.ascii_letters + string.digits
		return "".join(secrets.choice(alphabet) for _ in range(length))

	@staticmethod
	def _default_username(email: str) -> str:
		return email.split("@", 1)[0]

	@staticmethod
	def _normalize_role(role: str) -> str:
		value = str(role).strip().lower()
		if value in {"admin", "administrator"}:
			return "admin"
		if value in {"journaliste", "journalist", "editor", "viewer", "user"}:
			return "journaliste"
		raise HTTPException(
			status_code=status.HTTP_400_BAD_REQUEST,
			detail=f"Invalid role '{role}'. Allowed values: admin, journaliste.",
		)

	@staticmethod
	def _find_header_row(rows) -> tuple[int | None, dict[str, int]]:
		for index, row in enumerate(rows[:10]):
			headers = UserService._normalize_headers(row)
			if {"name", "email", "role"}.issubset(headers.keys()):
				return index, headers
		return None, {}

	@staticmethod
	def _normalize_headers(header_row) -> dict[str, int]:
		headers: dict[str, int] = {}
		for index, value in enumerate(header_row):
			if value is None:
				continue
			key = str(value).strip().lower()
			key_normalized = key.replace("é", "e").replace("è", "e").replace("ê", "e").replace("ô", "o")
			if key_normalized in {"name", "username", "nom"}:
				headers.setdefault("name", index)
			elif key_normalized in {"email", "courriel", "mail"}:
				headers.setdefault("email", index)
			elif key_normalized.startswith("role") or key_normalized.startswith("rôle") or key_normalized == "role":
				headers.setdefault("role", index)
			elif key_normalized in {"password", "mot de passe", "motdepasse"}:
				headers.setdefault("password", index)
			elif key_normalized in {"is_active", "active", "statut", "etat"}:
				headers.setdefault("is_active", index)
		return headers

	@staticmethod
	def _extract_row_data(headers: dict[str, int], row) -> dict[str, object]:
		return {field: row[index] if index < len(row) else None for field, index in headers.items()}

	@staticmethod
	def _is_row_empty(row) -> bool:
		return all(cell is None or str(cell).strip() == "" for cell in row)

	@staticmethod
	def _parse_bool(value, default: bool = True) -> bool:
		if value is None:
			return default
		if isinstance(value, bool):
			return value
		text = str(value).strip().lower()
		return text in {"1", "true", "yes", "on", "actif", "active"}

	@staticmethod
	def _validate_row(row_data: dict[str, object], row_number: int) -> list[str]:
		errors: list[str] = []
		name = str(row_data.get("name") or "").strip()
		email = str(row_data.get("email") or "").strip()
		role = str(row_data.get("role") or "").strip()
		if not name:
			errors.append(f"Row {row_number}: name is required.")
		if not email:
			errors.append(f"Row {row_number}: email is required.")
		if not role:
			errors.append(f"Row {row_number}: role is required.")
		else:
			try:
				UserService._normalize_role(role)
			except HTTPException as exc:
				errors.append(f"Row {row_number}: {exc.detail}")
		return errors


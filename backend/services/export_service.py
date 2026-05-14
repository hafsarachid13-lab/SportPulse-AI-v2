"""
Export Service — Person 5 Module (Upgraded)
Multi-format export system:
  • PDF — professional magazine-quality report with design system
  • Excel — multi-sheet workbook with styled headers
  • CSV — flat tabular export
  • JSON — full structured data
  • ZIP — all formats bundled
  • Export history tracking
"""

import os
import logging
import csv
import json
import uuid
import zipfile
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.services.pdf_designer import ReviewPDF

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────
# EXPORT SERVICE
# ──────────────────────────────────────────────────────────

class ExportService:
    """Reusable multi-format export service with history tracking."""

    _instance = None

    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self, output_dir: str = "static_exports"):
        if self._initialized:
            return
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

        # Font path
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.font_path = os.path.abspath(
            os.path.join(self.base_dir, "..", "fonts", "DejaVuSans.ttf")
        )

        # Export history
        self._export_history: List[Dict[str, Any]] = []
        self._initialized = True

    # ── Path helpers ──────────────────────────────────────

    def _get_filepath(self, filename: str) -> str:
        if "/" in filename or "\\" in filename or ".." in filename:
            raise ValueError(f"Invalid filename: {filename}")
        return os.path.join(self.output_dir, filename)

    def get_export_path(self, filename: str) -> str:
        return self._get_filepath(filename)

    def _track_export(self, fmt: str, filename: str) -> Dict[str, Any]:
        """Track export in history and return record."""
        filepath = self._get_filepath(filename)
        file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0

        record = {
            "id": str(uuid.uuid4()),
            "format": fmt,
            "filename": filename,
            "generated_at": datetime.now().isoformat(),
            "file_size": file_size,
        }
        self._export_history.insert(0, record)
        # Keep last 100 export records
        self._export_history = self._export_history[:100]
        return record

    # ── Export History ─────────────────────────────────────

    def get_export_history(self) -> List[Dict[str, Any]]:
        """Get export history records."""
        return self._export_history

    def get_export_stats(self) -> Dict[str, Any]:
        """Get export usage statistics."""
        from collections import Counter
        fmt_count = Counter(e["format"] for e in self._export_history)
        return {
            "total_exports": len(self._export_history),
            "by_format": dict(fmt_count),
            "recent_exports": self._export_history[:5],
        }

    def list_available_files(self) -> List[Dict[str, Any]]:
        """List all files in the export directory."""
        files = []
        if os.path.exists(self.output_dir):
            for fname in os.listdir(self.output_dir):
                fpath = os.path.join(self.output_dir, fname)
                if os.path.isfile(fpath):
                    files.append({
                        "filename": fname,
                        "size": os.path.getsize(fpath),
                        "modified_at": datetime.fromtimestamp(
                            os.path.getmtime(fpath)
                        ).isoformat(),
                    })
        return sorted(files, key=lambda x: x["modified_at"], reverse=True)

    # ──────────────────────────────────────────────────────
    # PDF EXPORT  (Professional Magazine-Quality Design)
    # ──────────────────────────────────────────────────────

    def generate_pdf(self, review_data: dict, filename: str = "review.pdf") -> str:
        """Generate a professionally formatted PDF report with magazine-quality design."""
        # Font setup
        body_font = "Helvetica"
        if os.path.exists(self.font_path):
            body_font = "DejaVu"

        date_str = review_data.get("generated_at", review_data.get("date", ""))[:10]
        metadata = review_data.get("metadata", {})
        sections = review_data.get("sections", {})

        # Create PDF with design system
        pdf = ReviewPDF(body_font=body_font, report_date=date_str)

        if body_font == "DejaVu":
            try:
                # Register the same TTF for both regular and bold so every
                # set_font("DejaVu", "B", ...) call works without crashing on
                # Arabic / non-latin Unicode characters.
                pdf.add_font("DejaVu", "",  self.font_path, uni=True)
                pdf.add_font("DejaVu", "B", self.font_path, uni=True)
            except TypeError:
                # Older fpdf2 builds don't accept the uni kwarg
                try:
                    pdf.add_font("DejaVu", "",  self.font_path)
                    pdf.add_font("DejaVu", "B", self.font_path)
                except Exception as e:
                    logger.warning(f"Could not load DejaVu font, falling back: {e}")
                    pdf.body_font = "Helvetica"
            except Exception as e:
                logger.warning(f"Could not load DejaVu font, falling back: {e}")
                pdf.body_font = "Helvetica"

        pdf.set_auto_page_break(auto=True, margin=15)

        # 1) Cover Page
        pdf.draw_cover(
            title=review_data.get("title", "Daily Sports Press Review"),
            date_str=date_str,
            metadata=metadata,
        )

        # 2) Executive Summary
        exec_summary = sections.get("executive_summary", {})
        pdf.draw_executive_summary(exec_summary)

        # 3) Top Headlines
        headlines = review_data.get(
            "highlights", review_data.get("top_headlines", [])
        )[:5]
        pdf.draw_headlines(headlines)

        # 4) Category Sections (one per sport)
        for sport, articles in review_data.get("categories", {}).items():
            if articles:
                pdf.draw_category(sport, articles)

        # 5) Trends & Insights
        trends = sections.get("trends", {})
        pdf.draw_trends(trends)

        # Output
        filepath = self._get_filepath(filename)
        pdf.output(filepath)
        logger.info(f"PDF exported: {filepath}")
        self._track_export("pdf", filename)
        return filename

    # ──────────────────────────────────────────────────────
    # EXCEL EXPORT
    # ──────────────────────────────────────────────────────

    def generate_excel(self, review_data: dict, filename: str = "review.xlsx") -> str:
        """Generate multi-sheet Excel workbook with styled headers."""
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="004E89")
        accent_fill = PatternFill("solid", fgColor="FF6B35")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        wrap_align = Alignment(wrap_text=True, vertical="top")
        center_align = Alignment(horizontal="center", vertical="center")

        def style_headers(ws, headers, fill=None):
            ws.append(headers)
            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = fill or header_fill
                cell.alignment = center_align
                cell.border = border

        # ── Sheet 1: Overview ──
        ws_overview = wb.active
        ws_overview.title = "Overview"
        ws_overview.column_dimensions["A"].width = 30
        ws_overview.column_dimensions["B"].width = 50

        ws_overview.append(["Sports Press Review"])
        ws_overview.merge_cells("A1:B1")
        ws_overview["A1"].font = Font(bold=True, size=16, color="004E89")

        date_str = review_data.get("generated_at", review_data.get("date", ""))[:10]
        ws_overview.append(["Date", date_str])
        ws_overview.append([])

        metadata = review_data.get("metadata", {})
        ws_overview.append(["Metric", "Value"])
        ws_overview["A4"].font = Font(bold=True)
        ws_overview["B4"].font = Font(bold=True)
        ws_overview.append(["Total Articles", metadata.get("total_articles", 0)])
        ws_overview.append(["Sources Count", metadata.get("sources_count", 0)])
        ws_overview.append(["Categories Count", metadata.get("categories_count", 0)])
        ws_overview.append([])

        ws_overview.append(["Top Headlines"])
        ws_overview[f"A{ws_overview.max_row}"].font = Font(bold=True, size=12)
        for headline in review_data.get("highlights", review_data.get("top_headlines", []))[:10]:
            ws_overview.append([headline])

        # ── Sheet 2: All Articles ──
        ws_all = wb.create_sheet(title="All Articles")
        all_headers = [
            "Title", "Source", "Sport", "Importance", "Credibility",
            "Published Date", "Summary",
        ]
        style_headers(ws_all, all_headers)

        col_widths = [50, 20, 15, 12, 12, 18, 60]
        for i, w in enumerate(col_widths, 1):
            ws_all.column_dimensions[get_column_letter(i)].width = w

        for row_idx, art in enumerate(review_data.get("articles", []), 2):
            ws_all.cell(row=row_idx, column=1, value=art.get("title", ""))
            ws_all.cell(row=row_idx, column=2, value=art.get("source", ""))
            ws_all.cell(row=row_idx, column=3, value=art.get("sport", ""))
            ws_all.cell(row=row_idx, column=4, value=float(art.get("importance_score", 0)))
            ws_all.cell(row=row_idx, column=5, value=float(art.get("credibility", 0.75)))
            ws_all.cell(row=row_idx, column=6, value=art.get("published_date", ""))
            summary_cell = ws_all.cell(row=row_idx, column=7, value=art.get("summary", ""))
            summary_cell.alignment = wrap_align
            for col in range(1, 8):
                ws_all.cell(row=row_idx, column=col).border = border

        # ── Per-sport sheets ──
        sport_headers = ["Title", "Source", "Importance Score", "Summary"]
        for sport, articles in review_data.get("categories", {}).items():
            safe_name = sport[:31].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=safe_name)
            style_headers(ws, sport_headers, fill=accent_fill)

            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 80

            for row_idx, art in enumerate(articles, 2):
                ws.cell(row=row_idx, column=1, value=art.get("title", ""))
                ws.cell(row=row_idx, column=2, value=art.get("source", ""))
                ws.cell(row=row_idx, column=3, value=float(art.get("importance_score", 0)))
                summary_cell = ws.cell(row=row_idx, column=4, value=art.get("summary", ""))
                summary_cell.alignment = wrap_align
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).border = border

        # ── Sources Sheet ──
        ws_sources = wb.create_sheet(title="Sources")
        source_headers = ["Source", "Articles"]
        style_headers(ws_sources, source_headers)
        ws_sources.column_dimensions["A"].width = 30

        sections = review_data.get("sections", {})
        top_sources = sections.get("top_sources", {}).get("sources", [])
        for row_idx, src in enumerate(top_sources, 2):
            ws_sources.cell(row=row_idx, column=1, value=src.get("source", ""))
            ws_sources.cell(row=row_idx, column=2, value=src.get("article_count", 0))

        filepath = self._get_filepath(filename)
        wb.save(filepath)
        logger.info(f"Excel exported: {filepath}")
        self._track_export("excel", filename)
        return filename

    # ──────────────────────────────────────────────────────
    # CSV EXPORT
    # ──────────────────────────────────────────────────────

    def generate_csv(self, review_data: dict, filename: str = "review.csv") -> str:
        """Generate flat CSV export."""
        filepath = self._get_filepath(filename)
        with open(filepath, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Category", "Title", "Source", "Importance Score",
                "Credibility", "Published Date", "Keywords", "Summary",
            ])

            for sport, articles in review_data.get("categories", {}).items():
                for art in articles:
                    keywords = art.get("keywords", [])
                    kw_str = "; ".join(keywords) if isinstance(keywords, list) else str(keywords)
                    writer.writerow([
                        sport,
                        art.get("title", ""),
                        art.get("source", ""),
                        float(art.get("importance_score", 0)),
                        float(art.get("credibility", 0.75)),
                        art.get("published_date", ""),
                        kw_str,
                        art.get("summary", ""),
                    ])
        logger.info(f"CSV exported: {filepath}")
        self._track_export("csv", filename)
        return filename

    # ──────────────────────────────────────────────────────
    # JSON EXPORT
    # ──────────────────────────────────────────────────────

    def generate_json(self, review_data: dict, filename: str = "review.json") -> str:
        """Generate full structured JSON export."""
        filepath = self._get_filepath(filename)
        with open(filepath, mode="w", encoding="utf-8") as f:
            json.dump(review_data, f, indent=4, ensure_ascii=False, default=str)
        logger.info(f"JSON exported: {filepath}")
        self._track_export("json", filename)
        return filename

    # ──────────────────────────────────────────────────────
    # ZIP EXPORT
    # ──────────────────────────────────────────────────────

    def generate_zip(
        self, review_data: dict, filename: str = "review_exports.zip"
    ) -> str:
        """Generate ZIP bundle with all format exports."""
        pdf_file = self.generate_pdf(review_data, "review.pdf")
        excel_file = self.generate_excel(review_data, "review.xlsx")
        csv_file = self.generate_csv(review_data, "review.csv")
        json_file = self.generate_json(review_data, "review.json")

        filepath = self._get_filepath(filename)
        with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(self._get_filepath(pdf_file), arcname=pdf_file)
            zipf.write(self._get_filepath(excel_file), arcname=excel_file)
            zipf.write(self._get_filepath(csv_file), arcname=csv_file)
            zipf.write(self._get_filepath(json_file), arcname=json_file)

        logger.info(f"ZIP exported: {filepath}")
        self._track_export("zip", filename)
        return filename